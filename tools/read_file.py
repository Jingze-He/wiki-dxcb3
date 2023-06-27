from openpyxl import load_workbook

'''
    解析Young大佬的单页式再生者图鉴，批量转化为可导入Wiki的行式再生者图鉴
'''
def parse_young_zaishengzhe(s):

    #Before
    res = []
    pifu = 13

    for i in range(5,30):
        if s['A'+str(i)].value and "再生者简介" in s['A'+str(i)].value:
            jiyi = i
            break

    #称号
    if "\n" in s['A1'].value: res.append(s['A1'].value.split("\n")[0])
    else: res.append(s['A1'].value.split(" ")[0])

    #姓名
    if "\n" in s['A1'].value: res.append(s['A1'].value.split("\n")[1])
    else: res.append(s['A1'].value.split(" ")[1])

    #昵称
    res.append(s.title)

    #记忆
    res.append(s['A' + str(jiyi+1)].value.replace("\n", "、")[:-1])

    #皮肤
    if True:
        tmp = []
        for i in range(pifu+4, jiyi):
            tmp.append(s['A' + str(i)].value.replace("\n", "").replace("典藏","(典藏)"))
        res.append("、".join(tmp))

    #发布时间
    res.append("2022-10-20")

    #出场排名
    res.append("-")

    #评价内容
    res.append("-")

    #定位
    res.append(s['C1'].value.split(":")[1])

    #武器
    res.append(s['D1'].value.split("：")[1])

    #特点
    res.append("、".join(s['E4'].value.split(":")[1].split(" ")))

    #属性
    if "/" in s['E5'].value: res.append("、".join(s['E5'].value.split("：")[1].split("/")))
    else: res.append("、".join(s['E5'].value.split("：")[1].split(" ")))
    #生命
    res.append(str(s['B2'].value))

    #护甲
    res.append(str(s['B3'].value))

    #魔盾
    res.append(str(s['B4'].value))

    #力魔
    res.append(str(s['B5'].value))

    #行动间隔
    res.append(str(s['D2'].value))

    #力魔品格
    res.append(s['A6'].value.split("\n")[1].split("+")[1].replace("/","、"))

    #生命品格
    res.append(s['A6'].value.split("\n")[2].split("+")[1].replace("/","、"))

    #双盾品格
    res.append(s['A6'].value.split("\n")[3].split("+")[1].replace("/","、"))

    #被动名称
    res.append(s['B8'].value)

    #被动初始
    res.append(s['C8'].value)

    #被动满级
    res.append(s['D8'].value)

    #战吼名称
    res.append(s['B9'].value)

    #战吼初始
    res.append(s['C9'].value)

    #战吼满级
    res.append(s['D9'].value)

    #战吼特效
    flag = True
    final_temp = []
    tmp = ""
    if s['E9'].value:
        for i in s['E9'].value.split("\n"):
            if flag:
                tmp = i
                flag = False
            else:
                flag = True
                final_temp.append(tmp + "：" + i)
    res.append("；".join(final_temp))

    #技能一名称
    res.append(s['B12'].value)

    #技能一初始
    res.append(s['C12'].value)

    #技能一满级
    res.append(s['D12'].value)

    #技能一特效
    flag = True
    final_temp = []
    tmp = ""
    if s['E12'].value:
        for i in s['E12'].value.split("\n"):
            if flag:
                tmp = i
                flag = False
            else:
                flag = True
                final_temp.append(tmp + "：" + i)
    res.append("；".join(final_temp))

    #技能二名称
    res.append(s['B10'].value)

    #技能二初始
    res.append(s['C10'].value)

    #技能二满级
    res.append(s['D10'].value)

    #技能二特效
    flag = True
    final_temp = []
    tmp = ""
    if s['E10'].value:
        for i in s['E10'].value.split("\n"):
            if flag:
                tmp = i
                flag = False
            else:
                flag = True
                final_temp.append(tmp + "：" + i)
    res.append("；".join(final_temp))

    #技能三名称
    res.append(s['B11'].value)

    #技能三初始
    res.append(s['C11'].value)

    #技能三满级
    res.append(s['D11'].value)

    #技能三特效
    flag = True
    final_temp = []
    tmp = ""
    if s['E11'].value:
        for i in s['E11'].value.split("\n"):
            if flag:
                tmp = i
                flag = False
            else:
                flag = True
                final_temp.append(tmp + "：" + i)
    res.append("；".join(final_temp))

    for i in range(len(res)):
        res[i] = res[i].replace("\n", "").replace(",", "，").replace(" ", "").replace("(", "（").replace(")", "）")
        if res[i] == "": res[i] = "-"
    print(" ".join(res))
    # for i in res:
    #     print(i)


wb = load_workbook('再生者资料.xlsx')
for i in wb.sheetnames:
    parse_young_zaishengzhe(wb[i])